"""Excel de produtos: exportação gera o arquivo, importação casa por
id/código de barras/SKU, cria produto novo quando não casa, nunca mexe em
estoque_atual, e some com a linha se a prévia (confirmar=False) for pedida."""

from decimal import Decimal
from io import BytesIO

import openpyxl
from sqlalchemy import select

from app.models.produto import Produto, Categoria
from app.services.produto_excel_service import gerar_excel_produtos, importar_excel_produtos
from tests.factories import criar_usuario, criar_produto


def _linhas(wb_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def _planilha(linhas: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "id", "codigo_barras", "sku", "nome", "fabricante", "descricao",
        "categoria", "fornecedor", "unidade_medida",
        "preco_custo", "preco_venda", "estoque_atual (não editável)",
        "estoque_minimo", "estoque_maximo", "ativo",
    ]
    ws.append(headers)
    for linha in linhas:
        ws.append(linha)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_exportar_traz_o_produto_cadastrado(db):
    p = await criar_produto(db, preco_venda="12.50", estoque="30", nome="Pão de Queijo")
    p.fabricante = "Casa do Pão"
    await db.flush()

    conteudo = await gerar_excel_produtos(db)
    linhas = _linhas(conteudo)

    assert len(linhas) == 1
    assert linhas[0]["nome"] == "Pão de Queijo"
    assert linhas[0]["fabricante"] == "Casa do Pão"
    assert linhas[0]["preco_venda"] == 12.50


async def test_importar_atualiza_por_codigo_de_barras_sem_mexer_no_estoque(db):
    u = await criar_usuario(db, perfil="admin")
    p = await criar_produto(db, preco_venda="10.00", estoque="42.000", nome="Pão Francês")
    p.codigo_barras = "789000000001"
    await db.flush()

    conteudo = _planilha([
        ["", "789000000001", None, "Pão Francês (novo nome)", "Marca X", None,
         "Padaria", None, "un", 3.0, 15.00, 999, 5, None, "sim"],
    ])

    resumo = await importar_excel_produtos(db, conteudo, confirmar=True, usuario_id=u.id)

    assert resumo.aplicado is True
    assert resumo.atualizados == 1
    assert resumo.criados == 0
    assert resumo.com_erro == 0

    await db.refresh(p)
    assert p.nome == "Pão Francês (novo nome)"
    assert p.fabricante == "Marca X"
    assert p.preco_venda == Decimal("15.00")
    assert p.estoque_atual == Decimal("42.000")  # intocado, mesmo a planilha dizendo 999


async def test_importar_cria_produto_novo_quando_nao_casa(db):
    u = await criar_usuario(db, perfil="admin")
    cat = Categoria(nome="Bebidas")
    db.add(cat)
    await db.flush()

    conteudo = _planilha([
        [None, "789000000099", "SKU-99", "Suco de Laranja 1L", "Fabricante Y", "Suco natural",
         "Bebidas", None, "l", 4.0, 9.90, 0, 2, None, "sim"],
    ])

    resumo = await importar_excel_produtos(db, conteudo, confirmar=True, usuario_id=u.id)

    assert resumo.criados == 1
    assert resumo.com_erro == 0

    criado = (await db.execute(
        select(Produto).where(Produto.codigo_barras == "789000000099")
    )).scalar_one()
    assert criado.nome == "Suco de Laranja 1L"
    assert criado.estoque_atual == Decimal("0.000")  # produto novo sempre nasce com zero
    assert criado.categoria_id == cat.id


async def test_importar_categoria_inexistente_vira_erro_e_nao_cria_nada(db):
    u = await criar_usuario(db, perfil="admin")
    conteudo = _planilha([
        [None, None, None, "Produto Fantasma", None, None,
         "Categoria Que Não Existe", None, "un", 1.0, 5.00, 0, 0, None, "sim"],
    ])

    resumo = await importar_excel_produtos(db, conteudo, confirmar=True, usuario_id=u.id)

    assert resumo.com_erro == 1
    assert resumo.criados == 0
    restantes = (await db.execute(select(Produto))).scalars().all()
    assert restantes == []


async def test_previa_nao_aplica_nada(db):
    u = await criar_usuario(db, perfil="admin")
    conteudo = _planilha([
        [None, "789000000050", None, "Produto Só de Prévia", None, None,
         "Padaria", None, "un", 1.0, 7.00, 0, 0, None, "sim"],
    ])
    # Cria a categoria "Padaria" pra não sobrar erro nessa linha
    db.add(Categoria(nome="Padaria"))
    await db.flush()

    resumo = await importar_excel_produtos(db, conteudo, confirmar=False, usuario_id=u.id)

    assert resumo.aplicado is False
    assert resumo.criados == 1  # reporta o que FARIA
    restantes = (await db.execute(
        select(Produto).where(Produto.codigo_barras == "789000000050")
    )).scalars().all()
    assert restantes == []  # mas não aplicou de verdade
