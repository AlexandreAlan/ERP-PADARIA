"""
ProdutoExcelService — Exporta o catálogo de produtos para Excel e importa
de volta com atualização inteligente (casa por id/código de barras/SKU).

Regra de ouro: a importação NUNCA mexe em estoque_atual. Quantidade em
estoque só muda por compra (NF-e/manual) ou ajuste de estoque — isso aqui é
só cadastro (nome, fabricante, categoria, fornecedor, preços, mínimos etc.),
pra não furar o rastro de auditoria que o resto do sistema mantém.
"""

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.produto import Produto, Categoria, Fornecedor

COLUNAS = [
    "id", "codigo_barras", "sku", "nome", "fabricante", "descricao",
    "categoria", "fornecedor", "unidade_medida",
    "preco_custo", "preco_venda", "estoque_atual (não editável)",
    "estoque_minimo", "estoque_maximo", "ativo",
]
_UNIDADES_VALIDAS = {"un", "kg", "g", "l", "ml", "pct"}


async def gerar_excel_produtos(db: AsyncSession, apenas_ativos: bool = True) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    stmt = select(Produto)
    if apenas_ativos:
        stmt = stmt.where(Produto.ativo.is_(True))
    produtos = (await db.execute(stmt.order_by(Produto.nome))).scalars().all()

    cats = {c.id: c.nome for c in (await db.execute(select(Categoria))).scalars().all()}
    forns = {f.id: f.razao_social for f in (await db.execute(select(Fornecedor))).scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(COLUNAS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r, p in enumerate(produtos, 2):
        valores = [
            p.id, p.codigo_barras, p.sku, p.nome, p.fabricante, p.descricao,
            cats.get(p.categoria_id, ""), forns.get(p.fornecedor_id, "") if p.fornecedor_id else "",
            p.unidade_medida, float(p.preco_custo), float(p.preco_venda),
            float(p.estoque_atual), float(p.estoque_minimo),
            float(p.estoque_maximo) if p.estoque_maximo is not None else None,
            "sim" if p.ativo else "não",
        ]
        for col, v in enumerate(valores, 1):
            ws.cell(row=r, column=col, value=v).border = border
        if r % 2 == 0:
            for col in range(1, len(COLUNAS) + 1):
                ws.cell(row=r, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    col_widths = [8, 16, 14, 30, 18, 30, 18, 22, 12, 12, 12, 20, 14, 14, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass
class LinhaResultado:
    linha: int
    nome: str
    acao: str            # criado | atualizado | erro
    motivo: Optional[str] = None


@dataclass
class ResumoImportacao:
    total_linhas: int = 0
    criados: int = 0
    atualizados: int = 0
    com_erro: int = 0
    aplicado: bool = False
    detalhes: list[LinhaResultado] = field(default_factory=list)


def _dec(valor, padrao: Optional[Decimal] = None) -> Optional[Decimal]:
    if valor is None or valor == "":
        return padrao
    try:
        return Decimal(str(valor))
    except InvalidOperation:
        return None


async def importar_excel_produtos(
    db: AsyncSession,
    conteudo: bytes,
    confirmar: bool,
    usuario_id: int,
) -> ResumoImportacao:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    cats = {c.nome.strip().lower(): c for c in (await db.execute(select(Categoria))).scalars().all()}
    forns = {f.razao_social.strip().lower(): f for f in (await db.execute(select(Fornecedor))).scalars().all()}

    resumo = ResumoImportacao()
    now = datetime.utcnow()

    linhas = list(ws.iter_rows(min_row=2, values_only=True))
    for idx, linha in enumerate(linhas, start=2):
        if linha is None or all(v is None for v in linha):
            continue
        resumo.total_linhas += 1

        vals = list(linha) + [None] * (len(COLUNAS) - len(linha))
        (raw_id, codigo_barras, sku, nome, fabricante, descricao,
         categoria_nome, fornecedor_nome, unidade_medida,
         preco_custo, preco_venda, _estoque_atual_ignorado,
         estoque_minimo, estoque_maximo, ativo_raw) = vals[:len(COLUNAS)]

        nome = (str(nome).strip() if nome is not None else "")
        if not nome:
            resumo.com_erro += 1
            resumo.detalhes.append(LinhaResultado(idx, "(sem nome)", "erro", "Nome é obrigatório"))
            continue

        codigo_barras = str(codigo_barras).strip() if codigo_barras not in (None, "") else None
        sku = str(sku).strip() if sku not in (None, "") else None

        # ── Localiza o produto: id > código de barras > SKU ──────────────
        produto: Optional[Produto] = None
        if raw_id not in (None, ""):
            try:
                produto = (await db.execute(
                    select(Produto).where(Produto.id == int(raw_id))
                )).scalar_one_or_none()
            except (ValueError, TypeError):
                produto = None
        if produto is None and codigo_barras:
            produto = (await db.execute(
                select(Produto).where(Produto.codigo_barras == codigo_barras)
            )).scalar_one_or_none()
        if produto is None and sku:
            produto = (await db.execute(select(Produto).where(Produto.sku == sku))).scalar_one_or_none()

        # ── Resolve categoria/fornecedor por nome ────────────────────────
        categoria = cats.get(str(categoria_nome).strip().lower()) if categoria_nome else None
        if categoria_nome and categoria is None:
            resumo.com_erro += 1
            resumo.detalhes.append(LinhaResultado(
                idx, nome, "erro", f"Categoria \"{categoria_nome}\" não existe — crie-a antes de importar."
            ))
            continue
        if categoria is None and produto is not None:
            categoria = None  # mantém a categoria atual do produto

        fornecedor = forns.get(str(fornecedor_nome).strip().lower()) if fornecedor_nome else None
        if fornecedor_nome and fornecedor is None:
            resumo.com_erro += 1
            resumo.detalhes.append(LinhaResultado(
                idx, nome, "erro", f"Fornecedor \"{fornecedor_nome}\" não existe — crie-o antes de importar."
            ))
            continue

        unidade = str(unidade_medida).strip().lower() if unidade_medida else (produto.unidade_medida if produto else "un")
        if unidade not in _UNIDADES_VALIDAS:
            resumo.com_erro += 1
            resumo.detalhes.append(LinhaResultado(
                idx, nome, "erro", f"Unidade \"{unidade_medida}\" inválida (use un/kg/g/l/ml/pct)."
            ))
            continue

        p_custo = _dec(preco_custo, produto.preco_custo if produto else Decimal("0.00"))
        p_venda = _dec(preco_venda, produto.preco_venda if produto else None)
        if p_venda is None or p_venda <= 0:
            resumo.com_erro += 1
            resumo.detalhes.append(LinhaResultado(idx, nome, "erro", "Preço de venda inválido ou ausente."))
            continue

        e_min = _dec(estoque_minimo, produto.estoque_minimo if produto else Decimal("0.000"))
        e_max = _dec(estoque_maximo, produto.estoque_maximo if produto else None)
        ativo = (produto.ativo if produto else True) if ativo_raw is None else \
            str(ativo_raw).strip().lower() in ("sim", "true", "1", "s", "yes")

        if produto is None:
            # Produto novo: categoria é obrigatória.
            if categoria is None:
                resumo.com_erro += 1
                resumo.detalhes.append(LinhaResultado(idx, nome, "erro", "Categoria é obrigatória para produto novo."))
                continue
            if confirmar:
                produto = Produto(
                    codigo_barras=codigo_barras, sku=sku, nome=nome,
                    fabricante=(str(fabricante).strip() if fabricante else None),
                    descricao=(str(descricao).strip() if descricao else None),
                    categoria_id=categoria.id,
                    fornecedor_id=fornecedor.id if fornecedor else None,
                    unidade_medida=unidade, preco_custo=p_custo, preco_venda=p_venda,
                    estoque_atual=Decimal("0.000"), estoque_minimo=e_min, estoque_maximo=e_max,
                    ativo=ativo, created_at=now, updated_at=now,
                )
                db.add(produto)
            resumo.criados += 1
            resumo.detalhes.append(LinhaResultado(idx, nome, "criado"))
        else:
            if confirmar:
                produto.codigo_barras = codigo_barras or produto.codigo_barras
                produto.sku = sku or produto.sku
                produto.nome = nome
                produto.fabricante = str(fabricante).strip() if fabricante else produto.fabricante
                produto.descricao = str(descricao).strip() if descricao else produto.descricao
                if categoria is not None:
                    produto.categoria_id = categoria.id
                if fornecedor_nome:
                    produto.fornecedor_id = fornecedor.id if fornecedor else produto.fornecedor_id
                produto.unidade_medida = unidade
                produto.preco_custo = p_custo
                produto.preco_venda = p_venda
                produto.estoque_minimo = e_min
                produto.estoque_maximo = e_max
                produto.ativo = ativo
                produto.updated_at = now
                # estoque_atual: propositalmente intocado.
            resumo.atualizados += 1
            resumo.detalhes.append(LinhaResultado(idx, nome, "atualizado"))

    if confirmar:
        await db.flush()
    else:
        await db.rollback()

    resumo.aplicado = confirmar
    return resumo
