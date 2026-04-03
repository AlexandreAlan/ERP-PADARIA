"""
RelatorioService — Gera relatórios em PDF (ReportLab) e Excel (openpyxl).
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Optional

from sqlalchemy import select, and_, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.venda import Venda, ItemVenda, Pagamento
from app.models.produto import Produto
from app.models.caixa import SessaoCaixa, Caixa
from app.models.usuario import Usuario
from app.config import get_settings

settings = get_settings()


async def _buscar_vendas(
    data_inicio: date,
    data_fim: date,
    db: AsyncSession,
    caixa_id: Optional[int] = None,
) -> list:
    dt_inicio = datetime.combine(data_inicio, datetime.min.time())
    dt_fim = datetime.combine(data_fim, datetime.max.time())

    q = (
        select(
            Venda.id,
            Venda.uuid,
            Venda.status,
            Venda.subtotal,
            Venda.desconto_valor,
            Venda.total,
            Venda.created_at,
            Usuario.nome.label("operador"),
            Caixa.nome.label("caixa"),
        )
        .join(Usuario, Usuario.id == Venda.usuario_id)
        .join(SessaoCaixa, SessaoCaixa.id == Venda.sessao_id)
        .join(Caixa, Caixa.id == SessaoCaixa.caixa_id)
        .where(
            Venda.status == "concluida",
            Venda.created_at.between(dt_inicio, dt_fim),
        )
        .order_by(Venda.created_at)
    )
    if caixa_id:
        q = q.where(SessaoCaixa.caixa_id == caixa_id)

    result = await db.execute(q)
    return result.all()


async def gerar_pdf_vendas(
    data_inicio: date,
    data_fim: date,
    db: AsyncSession,
    caixa_id: Optional[int] = None,
) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    vendas = await _buscar_vendas(data_inicio, data_fim, db, caixa_id)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Cabeçalho
    story.append(Paragraph(f"<b>{settings.padaria_nome}</b>", styles["Title"]))
    story.append(Paragraph(f"Relatório de Vendas: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    # Tabela
    headers = ["#", "Data/Hora", "Operador", "Caixa", "Subtotal", "Desconto", "Total"]
    table_data = [headers]

    total_geral = Decimal("0.00")
    for row in vendas:
        table_data.append([
            str(row.id),
            row.created_at.strftime("%d/%m/%Y %H:%M"),
            row.operador,
            row.caixa,
            f"R$ {row.subtotal:.2f}",
            f"R$ {row.desconto_valor:.2f}",
            f"R$ {row.total:.2f}",
        ])
        total_geral += Decimal(str(row.total))

    # Linha de total
    table_data.append(["", "", "", "TOTAL GERAL", "", "", f"R$ {total_geral:.2f}"])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f5f5f5")]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Total de vendas: {len(vendas)} | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


async def gerar_excel_vendas(
    data_inicio: date,
    data_fim: date,
    db: AsyncSession,
    caixa_id: Optional[int] = None,
) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    vendas = await _buscar_vendas(data_inicio, data_fim, db, caixa_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas"

    # Estilo cabeçalho
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["ID", "Data/Hora", "Operador", "Caixa", "Subtotal", "Desconto", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    total_geral = Decimal("0.00")
    for r, row in enumerate(vendas, 2):
        ws.cell(row=r, column=1, value=row.id)
        ws.cell(row=r, column=2, value=row.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=r, column=3, value=row.operador)
        ws.cell(row=r, column=4, value=row.caixa)
        ws.cell(row=r, column=5, value=float(row.subtotal))
        ws.cell(row=r, column=6, value=float(row.desconto_valor))
        ws.cell(row=r, column=7, value=float(row.total))
        total_geral += Decimal(str(row.total))

        if r % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border

    # Total
    last_row = len(vendas) + 2
    ws.cell(row=last_row, column=6, value="TOTAL GERAL").font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=float(total_geral)).font = Font(bold=True)

    # Largura das colunas
    col_widths = [8, 18, 20, 15, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Formata colunas monetárias
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = '"R$"#,##0.00'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
